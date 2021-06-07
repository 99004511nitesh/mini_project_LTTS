import openpyxl


def actions():

    data_1 = "pyexcel.xlsx"
    info_1 = openpyxl.load_workbook(data_1)

    sh_1 = info_1["semester marks"]
    for name_1 in sh_1['A']:
        print(name_1.value)

    array_1 = {}
    columns = ['PS NO']

    for sheet in info_1.worksheets:
        for i in range(2, sheet.max_row+1):
            for j in range(2, sheet.max_column+1):
                if sheet.cell(row=i, column=1).value not in array_1:
                    array_1[sheet.cell(row=i, column=1).value] = []
                if sheet.cell(row=1, column=j).value not in columns:
                    columns.append(sheet.cell(row=1, column=j).value)
                array_1[sheet.cell(row=i, column=1).value].append(sheet.cell(row=i, column=j).value)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Total Info"

    for i in range(len(columns)):
        sheet.cell(row=1, column=i+1).value = columns[i]
    var_k = [int(input('Select a PS NO: '))]
    for i in range(2, len(var_k)+2):
        sheet.cell(row=i, column=1).value = var_k[i-2]
        c_var = 2
    for j in array_1[var_k[i-2]]:
        sheet.cell(row=i, column=c_var).value = j
        c_var += 1
    workbook.save("Output.xlsx")


actions()